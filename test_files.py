import time
import os
import zipfile
import csv

from selene.support.shared import browser
from pathlib import Path
from conftest import DOWNLOAD_DIR
from selene import command
from pypdf import PdfReader
from openpyxl import load_workbook

def test_files():
    filename_pdf = "file-sample_150kB.pdf"
    filename_xlsx = "file_example_XLSX_10.xlsx"
    filename_csv = "file_example_CSV_5000.csv"

    browser.open("/sample-xls-download/")
    browser.element("button.fc-button.fc-cta-do-not-consent").click()
    browser.element(f"a[href*='{filename_xlsx}']").click()
    filepath = Path(DOWNLOAD_DIR) / filename_xlsx
    browser.wait.until(lambda: filepath.exists())

    browser.open("/sample-pdf-download/")
    browser.element(f"a[href*='{filename_pdf}']").perform(command.js.click)
    filepath = Path(DOWNLOAD_DIR) / filename_pdf
    browser.wait.until(lambda: filepath.exists())

    browser.open("/text-files-and-archives-download/")
    browser.element(f"a[href*='{filename_csv}']").perform(command.js.click)
    filepath = Path(DOWNLOAD_DIR) / filename_csv
    browser.wait.until(lambda: filepath.exists())

    zip_path = os.path.join(DOWNLOAD_DIR, 'test_files.zip')

    with zipfile.ZipFile(zip_path, 'w') as zf:
        for file in (filename_pdf, filename_xlsx, filename_csv):
            add_file = os.path.join(DOWNLOAD_DIR, file)
            zf.write(add_file, os.path.basename(add_file))

    for file in (filename_pdf, filename_xlsx, filename_csv):
        os.remove(os.path.join(DOWNLOAD_DIR, file))

    with zipfile.ZipFile(zip_path) as zip_file:
        with zip_file.open(filename_pdf) as pdf_file:
            reader = PdfReader(pdf_file)
            number_of_pages = len(reader.pages)
            first_page_text = reader.pages[0].extract_text()
            assert number_of_pages == 4
            assert first_page_text.startswith("Lorem ipsum")

    with zipfile.ZipFile(zip_path) as zip_file:
        with zip_file.open(filename_csv) as csv_file:
            content = csv_file.read().decode('utf-8-sig')
            csvreader = list(csv.reader(content.splitlines()))
            second_row = csvreader[1]

            assert second_row[1] == "Dulce"
            assert second_row[2] == "Abril"


    with zipfile.ZipFile(zip_path) as zip_file:
        with zip_file.open(filename_xlsx) as xlsx_file:
            workbook = load_workbook(xlsx_file)
            sheet = workbook.active
            assert sheet.cell(row=2, column=2).value == "Dulce"
            assert sheet.cell(row=2, column=3).value == "Abril"
